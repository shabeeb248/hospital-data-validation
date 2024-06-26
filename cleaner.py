from openpyxl import load_workbook
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from fuzzywuzzy import process
import re

def convert(file):
    wb = load_workbook(file)
    sheet_names = wb.sheetnames
    dfdict = excel_to_dataframes(uploaded_file=file, sheetnames=sheet_names)
    cleaned_dfdict = validate_all(dfdict=dfdict)
    merged = newindex(dfdict=cleaned_dfdict)
    validate_columns = merged.columns[merged.columns.str.endswith('-VALIDATE')]

    # Flip boolean values in the selected columns
    merged[validate_columns] = ~merged[validate_columns]  
    json_data = merged.to_json(orient='records')  
    return json_data


def excel_to_dataframes(uploaded_file, sheetnames):
    
    dfs_dict = {}
    count = 0
    for sheet_name in sheetnames:
        if sheet_name=="MAIN MENU":
            continue
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)

        if count==0:
            print(df)
            count+=1
        # Select only the first 8 columns
        while len(df.columns) < 9:
            df[len(df.columns)] = np.nan
        
        # Select only the first 9 columns
        df = df.iloc[:, :9]

        # Remove rows with at least one NaN value
        df = df.dropna()
        
        df.columns = ['STATUS', 'DATE', 'SHIFT', 'HOURS', 'RATE', 'COST', 'ON CALL', 'ROLE', 'UNIT']
        df = df[df['STATUS'].isin(['NEW', 'CURRENT', 'VACANT', 'PENDING'])]
        
        #df['HOSPITAL'] = sheet_name
        df = df.reset_index(drop=True)
        for col in df.columns:
            df[f'{col}-VALIDATE'] = True
        df['SHIFT START'] = np.nan
        df["STATE"],df["HOSPITAL"] = get_state(df, sheet_name)
        df["CALCULATION-VALIDATE"] = True
        df['SHIFT END'] = np.nan
        dfs_dict[sheet_name] = df
    
    return dfs_dict


def get_state(data, sheetname):
    address_pattern = re.compile(r'address.*')
    address = ""
    for column in data.columns:
        for index, value in enumerate(data[column]):
            if isinstance(value, str):
                match = address_pattern.search(value)
                if match:
                    address = match.group()
                    address = address.split(":")[-1].strip()
                    print(f'Extracted address: {address}')
                    print(f'Column: {column}, Row: {index}')

    df = pd.read_csv("data/hospstate.csv")
    req = df[["hospital-name","state","geo-address"]]
    if address!="":
        bestmatchadd = process.extractOne(address, req["geo-address"])
        name_map = dict(zip(req['geo-address'], req['hospital-name']))
        state_map = dict(zip(req['hospital-name'], req['state']))
        return state_map[name_map[bestmatchadd[0]]],name_map[bestmatchadd[0]]
    else:
        bestmatchadd = process.extractOne(sheetname, req["hospital-name"])
        state_map = dict(zip(req['hospital-name'], req['state']))
        return state_map[bestmatchadd[0]],bestmatchadd[0]
    


# %%
def validate_date(df):
    # Initialize an empty list to store boolean values indicating whether each row is valid
    indices = []
    # Iterate through each row in the dataframe
    for index, row in df.iterrows():
        try:
            # Attempt to parse the date value using the specified format
            date = pd.to_datetime(row['DATE'])
            df.at[index, 'DATE'] = date.strftime('%Y-%m-%d')
            df.at[index, 'DATE-VALIDATE'] = True
        except ValueError:
            # If parsing fails, append False to the validation list
            df.at[index, 'DATE-VALIDATE'] = False
            indices.append(index)
            print(index, "date validate failed")
    
    return df, indices

# %%
def validate_shift(df):
    # Iterate through each row in the dataframe
    indices = []
    for index, row in df.iterrows():
        shift_value = row['SHIFT']
        
        # Clean the shift value and split it into start and end times
        try:
            shift_value = str(shift_value).replace(" ","")
            
            # Check if the shift value is in the correct format
            if len(shift_value) != 9:
                # If the shift value is not in the correct format, set shift-validate to False for this row
                
                df.at[index, 'SHIFT-VALIDATE'] = False
                indices.append(index)

            else:
                shift_start = shift_value[:4]
                shift_end = shift_value[5:]
                # If the shift value is in the correct format, set shift-validate to True for this row
                df.at[index, 'SHIFT-VALIDATE'] = True
                
                df.at[index, 'SHIFT START'] = datetime.strptime(shift_start, "%H%M").time()
                df.at[index, 'SHIFT END'] = datetime.strptime(shift_end, "%H%M").time()
                # Set the 'SHIFT START' and 'SHIFT END' values
                if df.at[index, 'SHIFT START']>=df.at[index, 'SHIFT END']:
                    date_obj = datetime.strptime(str(df.at[index, 'DATE']),"%Y-%m-%d")
                    # Step 2: Add one day to the datetime object
                    next_date_obj = date_obj + timedelta(days=1)

                    # If you need the next date as a string in the same format:
                    next_date_str = next_date_obj.strftime("%Y-%m-%d")

                    df.at[index, 'SHIFT END'] = next_date_str + "T" + str(df.at[index, 'SHIFT END'])+ ".000Z"
                else:
                    df.at[index, 'SHIFT END'] = str(df.at[index, 'DATE']) + "T" + str(df.at[index, 'SHIFT END']) + ".000Z"
                df.at[index, 'SHIFT START'] = str(df.at[index, 'DATE']) + "T" + str(df.at[index, 'SHIFT START']) + ".000Z"
                
            
        except Exception as e:
            df.at[index, 'SHIFT-VALIDATE'] = False
            indices.append(index)
        

    return df,indices

# %%
def validate_hours(df):
    indices = []
    # Iterate through each row in the dataframe
    for index, row in df.iterrows():
        shift_value = row['SHIFT']
        hours_value = row['HOURS']

        try:
        # Check if the shift value is in the correct format "HHMM-HHMM"
            if not isinstance(shift_value, str) or not len(shift_value) == 9 or not shift_value[4] == '-':
                # If the shift value is not in the correct format, set hours-validate to False for this row
                df.at[index, 'HOURS-VALIDATE'] = False
                indices.append(index)
            else:
                # If the shift value is in the correct format, calculate the hours worked
                start_hour, end_hour = shift_value.split('-')
                start_hour = int(start_hour[:2]) + int(start_hour[2:]) / 60
                end_hour = int(end_hour[:2]) + int(end_hour[2:]) / 60
                hours_worked = end_hour - start_hour
                if hours_worked<0:
                    hours_worked = 24 + hours_worked
                # Check if the calculated hours match the value in the HOURS column
                if hours_worked != hours_value:  # Allowing for small floating point differences
                    # If the calculated hours do not match, assign the correct value
                    #df.at[index, 'HOURS'] = hours_worked
                    indices.append(index)
                    # Set hours-validate to False for this row
                    df.at[index, 'CALCULATION-VALIDATE'] = False
                else:
                    # If the calculated hours match, set hours-validate to True for this row
                    df.at[index, 'HOURS-VALIDATE'] = True
        except:
            df.at[index, 'HOURS-VALIDATE'] = False
    
    return df, indices


# %%
def validate_rate(df):
    # Iterate through each row in the dataframe
    indices = []
    for index, row in df.iterrows():
        try:
            rate_value = row['RATE']
            cost_value = row['COST']
            rate = int(rate_value)
            # Check if rate equals cost and hours is not equal to 1
            if rate_value == cost_value or type(rate_value)==str:
                # If rate equals cost but hours is not 1, set rate-validate to False for this row
                indices.append(index)
                
                df.at[index, 'CALCULATION-VALIDATE'] = False
            else:
                # Otherwise, set rate-validate to True for this row
                df.at[index, 'RATE-VALIDATE'] = True
        except:
            indices.append(index)
            df.at[index, 'RATE-VALIDATE'] = False
    
    return df,indices

# %%
def validate_cost(df):
    indices = []
    for index, row in df.iterrows():
        try:
            rate_value = row['RATE']
            hours_value = row['HOURS']
            cost_value = row['COST']

            cost = int(cost_value)
            
            expected_cost = rate_value * hours_value
            if expected_cost != cost_value:
                indices.append(index)
                df.at[index, 'CALCULATION-VALIDATE'] = False
            else:
                df.at[index, 'COST-VALIDATE'] = True
        except:
            indices.append(index)
            df.at[index, 'COST-VALIDATE'] = False
    
    return df, indices

# %%
def validate_oncall(df):
    # Iterate through each row in the dataframe
    indices = []
    for index, row in df.iterrows():
        oncall_value = row['ON CALL']
        
        # Check if cost value is either 'yes' or 'no'
        try:
            if oncall_value.lower().strip() not in ['yes', 'no']:
                # If cost value is not 'yes' or 'no', set cost-validate to False for this row
                df.at[index, 'ON CALL-VALIDATE'] = False
                indices.append(index)
            else:
                # Otherwise, set cost-validate to True for this row
                df.at[index, 'ON CALL-VALIDATE'] = True
        except:
             df.at[index, 'ON CALL-VALIDATE'] = False
    
    return df, indices

# %%
def validate_roles(df):
    valid_roles = ['CMO Senior', 'REGISTRAR', 'RMO', 'SRMO', 'CMO NON IC', 'REGISTRAR IC']
    indices = []
    # Iterate through each row in the dataframe
    for index, row in df.iterrows():
        role_value = row['ROLE']
        try:
        # Check if role value is in the predefined list of valid roles
            if role_value.strip().upper() not in valid_roles:
                # If role value is not in the predefined list, set roles-validate to False for this row
                indices.append(index)
                df.at[index, 'ROLE-VALIDATE'] = False
            else:
                # Otherwise, set roles-validate to True for this row
                df.at[index, 'ROLE-VALIDATE'] = True
        except:
            df.at[index, 'ROLE-VALIDATE'] = False
    
    return df, indices

# %%


def validate_units(df):
    valid_units = ['ANAESTH', 'ED', 'FACILITY', 'ICU', 'MEDICAL', 'O & G', 'ONCOLOGY', 'ORTHO', 'PAEDS', 'PSYCH', 'SURGICAL', 'WARDS']
    indices = []
    # Iterate through each row in the dataframe
    for index, row in df.iterrows():
        unit_value = row['UNIT']
        try:
            # Check if unit value is in the predefined list of valid units
            if unit_value.strip().upper() not in valid_units:
                # If unit value is not in the predefined list, set unit-validate to False for this row
                indices.append(index)
                df.at[index, 'UNIT-VALIDATE'] = False
            else:
                # Otherwise, set unit-validate to True for this row
                df.at[index, 'UNIT-VALIDATE'] = True
        except:
            df.at[index, 'UNIT-VALIDATE'] = False
    
    return df, indices

validations = [validate_date, validate_hours, validate_rate, validate_cost, validate_hours, validate_roles, validate_oncall, validate_units, validate_shift]

def validate_all(dfdict):
    validated = dfdict
    failed = {}
    for val in validations:
        for x in dfdict.keys():
            validated[x],failed[x] = val(validated[x])
    return validated


def newindex(dfdict):
    dfs_list = []
    sheetname_abrs = []
# Iterate through each dataframe in the dictionary
    for sheet_name, df in dfdict.items():
        # Shorten the date and concatenate the sheet name with it
        sheetname_abr = "".join([x[0:3] for x in sheet_name.split(" ")])
        sheetname_abrs.append(sheetname_abr)
        df['SERIAL NO'] = sheetname_abr + pd.to_datetime(df['DATE']).dt.strftime('%y%m%d') + df['SHIFT'].astype(str).str.replace("-", "").replace(" ","")   
        # Append the modified dataframe to the list
        df["DATE"] = pd.to_datetime(df['DATE']).dt.strftime("%Y-%m-%d") + "T00:00:00.000Z"
        dfs_list.append(df)

    merged_df = pd.concat(dfs_list, ignore_index=True)
    if len(set(sheetname_abrs)) != len(sheetname_abrs):
        print("not unique")
    return merged_df




