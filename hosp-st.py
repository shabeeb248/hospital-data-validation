import base64
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.filterwarnings('ignore')
import numpy as np
import cleaner
import datetime

@st.cache_resource()
def load_data(file):
    wb = load_workbook(file)
    sheet_names = wb.sheetnames
    dfdict = cleaner.excel_to_dataframes(uploaded_file=file, sheetnames=sheet_names)
    cleaned_dfdict = cleaner.validate_all(dfdict=dfdict)
    merged = cleaner.newindex(dfdict=cleaned_dfdict)
    validate_columns = merged.columns[merged.columns.str.endswith('-VALIDATE')]
    print(validate_columns)
    # Flip boolean values in the selected columns
    merged[validate_columns] = ~merged[validate_columns]    
    return merged


def main():
    st.title("Excel File Upload App")
    st.write("last pushed May 21 1:49 AM")
    st.write("This app allows you to upload an Excel file and view its contents.")

    # File upload
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

    if uploaded_file is not None:
        # Load data using cache
        merged = load_data(uploaded_file)
        json_data = merged.to_json(orient='records')

        def download_json(data, filename):
            b64 = base64.b64encode(data.encode()).decode()
            href = f'<a href="data:application/json;base64,{b64}" download="{filename}">Download JSON file</a>'
            return href
        
        st.write(merged.iloc[30])
        st.markdown(download_json(json_data, 'data.json'), unsafe_allow_html=True)

        

if __name__ == "__main__":
    main()
